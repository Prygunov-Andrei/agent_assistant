#!/usr/bin/env python
"""
Скрипт для создания шаблона Excel файла для импорта персон
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_template():
    """Создает Excel шаблон для импорта персон"""
    
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Импорт персон'

    # Настройки стилей
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки столбцов
    headers = [
        'Тип персоны',
        'Фамилия',
        'Имя',
        'Телефоны',
        'Telegram',
        'Email',
        'Ссылка на Кинопоиск'
    ]

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Примеры данных
    examples = [
        ['Режиссер', 'Иванов', 'Иван', '+79001234567 +79009876543', '@ivanov', 'ivan@mail.ru ivan@work.com', 'https://www.kinopoisk.ru/name/12345/'],
        ['КД', 'Петрова', 'Мария', '+79101234567', '@petrova_casting', 'maria@gmail.com', ''],
        ['Продюсер', 'Сидоров', 'Петр', '', '@sidorov_prod', 'petr@company.ru', 'https://www.kinopoisk.ru/name/67890/'],
    ]

    # Добавляем примеры
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 50

    # Высота строки заголовка
    ws.row_dimensions[1].height = 40

    # Добавляем инструкции на отдельный лист
    ws_info = wb.create_sheet('Инструкция')
    ws_info.column_dimensions['A'].width = 100

    instructions = [
        ('ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ', 'header'),
        ('', 'normal'),
        ('1. Тип персоны (обязательно):', 'bold'),
        ('   Допустимые значения: Режиссер, КД (или Кастинг-директор), Продюсер', 'normal'),
        ('', 'normal'),
        ('2. Фамилия (обязательно):', 'bold'),
        ('   Фамилия персоны', 'normal'),
        ('', 'normal'),
        ('3. Имя (обязательно):', 'bold'),
        ('   Имя персоны', 'normal'),
        ('', 'normal'),
        ('4. Телефоны (опционально):', 'bold'),
        ('   Несколько номеров через пробел. Пример: +79001234567 +79009876543', 'normal'),
        ('', 'normal'),
        ('5. Telegram (опционально):', 'bold'),
        ('   Несколько никнеймов через пробел. Можно с @ или без. Пример: @ivanov @ivanov_official', 'normal'),
        ('', 'normal'),
        ('6. Email (опционально):', 'bold'),
        ('   Несколько адресов через пробел. Пример: ivan@mail.ru ivan.work@gmail.com', 'normal'),
        ('', 'normal'),
        ('7. Ссылка на Кинопоиск (опционально):', 'bold'),
        ('   Полная ссылка на профиль персоны. Пример: https://www.kinopoisk.ru/name/12345/', 'normal'),
        ('', 'normal'),
        ('ВАЖНО:', 'header'),
        ('• Не удаляйте строку с заголовками', 'normal'),
        ('• Максимум 5 контактов каждого типа', 'normal'),
        ('• Система автоматически проверит дубликаты', 'normal'),
        ('• При совпадении вы сможете выбрать: создать новую персону или обновить существующую', 'normal'),
    ]

    for row_num, (text, style) in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_num, column=1)
        cell.value = text
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        if style == 'header':
            cell.font = Font(bold=True, size=14, color='4472C4')
        elif style == 'bold':
            cell.font = Font(bold=True, size=11)
        else:
            cell.font = Font(size=10)

    # Определяем путь для сохранения
    current_dir = os.path.dirname(os.path.abspath(__file__))
    template_dir = os.path.join(current_dir, 'templates')
    
    # Создаем директорию если нужно
    os.makedirs(template_dir, exist_ok=True)
    
    # Сохраняем файл
    template_path = os.path.join(template_dir, 'person_import_template.xlsx')
    wb.save(template_path)
    
    print(f'Шаблон создан успешно: {template_path}')
    return template_path


if __name__ == '__main__':
    create_template()

