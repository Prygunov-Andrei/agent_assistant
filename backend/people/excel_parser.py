"""
Парсер Excel файлов для импорта персон
"""
import openpyxl
from typing import List, Dict, Optional
import re


class ExcelPersonParser:
    """Парсер XLSX файлов с персонами"""
    
    # Маппинг столбцов (номер столбца -> имя поля)
    COLUMN_MAPPING = {
        0: 'person_type',
        1: 'last_name',
        2: 'first_name',
        3: 'phones',
        4: 'telegrams',
        5: 'emails',
        6: 'kinopoisk_url',
    }
    
    # Маппинг типов персон из Excel в значения модели
    PERSON_TYPE_MAPPING = {
        'кд': 'casting_director',
        'кастинг-директор': 'casting_director',
        'кастинг директор': 'casting_director',
        'casting director': 'casting_director',
        'режиссер': 'director',
        'режиссёр': 'director',
        'director': 'director',
        'продюсер': 'producer',
        'producer': 'producer',
    }
    
    def parse_file(self, file_path: str) -> List[Dict]:
        """
        Парсит XLSX файл и возвращает список записей
        
        Args:
            file_path: Путь к файлу
            
        Returns:
            List[Dict]: Список записей с данными персон
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            records = []
            # Начинаем со второй строки (пропускаем заголовки)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                # Пропускаем пустые строки
                if self._is_empty_row(row):
                    continue
                    
                record = self._parse_row(row, row_idx)
                if record:
                    records.append(record)
            
            workbook.close()
            return records
            
        except Exception as e:
            raise ValueError(f"Ошибка при чтении файла: {str(e)}")
    
    def _is_empty_row(self, row) -> bool:
        """Проверяет, является ли строка пустой"""
        return all(cell.value is None or str(cell.value).strip() == '' for cell in row)
    
    def _parse_row(self, row, row_number: int) -> Optional[Dict]:
        """
        Парсит одну строку Excel
        
        Args:
            row: Строка из openpyxl
            row_number: Номер строки (для отслеживания)
            
        Returns:
            Dict: Данные персоны или None если строка пустая
        """
        try:
            # Извлекаем значения ячеек
            person_type_raw = self._get_cell_value(row, 0)
            last_name = self._get_cell_value(row, 1)
            first_name = self._get_cell_value(row, 2)
            phones_raw = self._get_cell_value(row, 3)
            telegrams_raw = self._get_cell_value(row, 4)
            emails_raw = self._get_cell_value(row, 5)
            kinopoisk_url = self._get_cell_value(row, 6)
            
            # Нормализуем тип персоны
            person_type = self._normalize_person_type(person_type_raw)
            
            # Парсим контакты
            phones = self._split_contacts(phones_raw)
            telegrams = self._split_contacts(telegrams_raw)
            emails = self._split_contacts(emails_raw)
            
            # Нормализуем telegram (убираем @ если есть)
            telegrams = [self._normalize_telegram(t) for t in telegrams]
            
            data = {
                'row_number': row_number,
                'data': {
                    'person_type': person_type,
                    'last_name': last_name.strip() if last_name else '',
                    'first_name': first_name.strip() if first_name else '',
                    'phones': phones,
                    'telegrams': telegrams,
                    'emails': emails,
                    'kinopoisk_url': kinopoisk_url.strip() if kinopoisk_url else None,
                },
                'validation_errors': [],
                'potential_duplicates': []
            }
            
            return data
            
        except Exception as e:
            return {
                'row_number': row_number,
                'data': {},
                'validation_errors': [f"Ошибка парсинга строки: {str(e)}"],
                'potential_duplicates': []
            }
    
    def _get_cell_value(self, row, col_index: int) -> Optional[str]:
        """Получает значение ячейки и конвертирует в строку"""
        try:
            cell = row[col_index]
            value = cell.value
            
            if value is None:
                return None
            
            # Конвертируем в строку и убираем лишние пробелы
            return str(value).strip()
            
        except IndexError:
            return None
    
    def _split_contacts(self, value: Optional[str]) -> List[str]:
        """
        Разделяет контакты через пробел или запятую
        
        Args:
            value: Строка с контактами
            
        Returns:
            List[str]: Список контактов
        """
        if not value:
            return []
        
        # Разделяем по пробелам и запятым
        contacts = re.split(r'[\s,;]+', str(value))
        
        # Убираем пустые значения и лишние пробелы
        contacts = [c.strip() for c in contacts if c.strip()]
        
        return contacts
    
    def _normalize_person_type(self, value: Optional[str]) -> str:
        """
        Нормализует тип персоны из различных вариантов написания
        
        Args:
            value: Сырое значение из Excel
            
        Returns:
            str: Нормализованный тип или пустая строка
        """
        if not value:
            return ''
        
        # Приводим к нижнему регистру для сравнения
        normalized = value.lower().strip()
        
        # Ищем в маппинге
        return self.PERSON_TYPE_MAPPING.get(normalized, value.strip())
    
    def _normalize_telegram(self, telegram: str) -> str:
        """
        Нормализует Telegram username (убирает @ если есть)
        
        Args:
            telegram: Telegram username
            
        Returns:
            str: Нормализованный username
        """
        telegram = telegram.strip()
        
        # Убираем @ в начале
        if telegram.startswith('@'):
            telegram = telegram[1:]
        
        # Убираем t.me/ если есть
        telegram = re.sub(r'^(?:https?://)?(?:t\.me/|telegram\.me/)', '', telegram)
        
        return telegram
    
    def validate_file(self, file_path: str) -> Dict:
        """
        Валидирует файл без полного парсинга
        
        Args:
            file_path: Путь к файлу
            
        Returns:
            Dict: Результат валидации
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Проверяем, что есть хотя бы заголовок и одна строка данных
            if sheet.max_row < 2:
                return {
                    'valid': False,
                    'error': 'Файл пуст или содержит только заголовки'
                }
            
            # Проверяем количество столбцов
            if sheet.max_column < 3:
                return {
                    'valid': False,
                    'error': 'Недостаточно столбцов в файле (минимум 3: тип, фамилия, имя)'
                }
            
            workbook.close()
            
            return {
                'valid': True,
                'rows_count': sheet.max_row - 1,  # Минус заголовок
                'columns_count': sheet.max_column
            }
            
        except Exception as e:
            return {
                'valid': False,
                'error': f'Ошибка чтения файла: {str(e)}'
            }

